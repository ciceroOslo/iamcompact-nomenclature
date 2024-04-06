"""For checking scenario and variable names used in the result files"""

# %%
# Imports
from pathlib import Path
import typing as tp
import pickle
import hashlib
import itertools

import pyam
import pandas as pd
import nomenclature
import openpyxl

import iamcompact_nomenclature as icnom
from iamcompact_nomenclature.aggregation import check_var_aggregates
from iamcompact_nomenclature import var_utils


# %%
# Create a helper function that takes a function as an argument, and returns
# a modified function which will return the return value of the original
# function if it returns, but catch any exceptions and return them. The function
# should be properly type-annotated, with ParamSpecs to ensure that the
# modified function has the same signature and return type as the original
# function, except that it can return an exception as well as the original
# return type.
PS = tp.ParamSpec('PS')  # Parameter specifiction
RT = tp.TypeVar('RT')  # Return type
def return_exceptions(
    func: tp.Callable[PS, RT]
) -> tp.Callable[PS, RT|Exception]:
    def wrapper(*args: PS.args, **kwargs: PS.kwargs) -> RT|Exception:
        try:
            return func(*args, **kwargs)
        except Exception as e:
            return e
    return wrapper

# %%
# Define a function to open IAMC excel files with multiple sheets
def open_multisheet_iamc(path: Path|str) -> \
        pyam.IamDataFrame|Exception|dict[str, pyam.IamDataFrame|Exception]:
    sheetnames: list[str] = openpyxl.open(path).sheetnames
    if len(sheetnames) == 1:
        return return_exceptions(pyam.IamDataFrame)(path)
    else:
        return {
            sheetname: return_exceptions(pyam.IamDataFrame) \
                (path, sheet_name=sheetname)
            for sheetname in sheetnames
        }

# %%
# Get data files, and create a nested dict of IamDataFrames
# If a cached pickle file exists, load it. Otherwise, load the data files and
# save the dict as a pickle file.
data_root: Path = Path.cwd() / 'study_results'
cache_file: Path = data_root / 'data_dict.pkl'
FORCE_RELOAD: bool = False
pickle_hash: str = 'e59a823d322e5b9fadf03fde3e014fa4'
write_cache: bool = True

if cache_file.exists() and not FORCE_RELOAD:
    # First check that the md5sum hash of the pickle file matches `pickle_hash`.
    # Raise an error if not.
    print(f'Reading cache file {cache_file}, expecting MD5 hash {pickle_hash}')
    with open(cache_file, 'rb') as f:
        cache_file_content: bytes = f.read()
        cache_file_hash = hashlib.md5(cache_file_content).hexdigest()
        if cache_file_hash != pickle_hash:
            raise ValueError(
                f'Hash of pickle file {cache_file} does not match expected hash'
            )
    # Load the pickle file using the content that was already read
    data_dict: dict[str, pyam.IamDataFrame | Exception |
                    dict[str, pyam.IamDataFrame | Exception]] = pickle.loads(
        cache_file_content
    )
    print('Successfully loaded cache file.')
else:
    if not cache_file.exists():
        print(f'Cache file {cache_file} does not exist. Loading data files.')
    elif FORCE_RELOAD:
        print(f'FORCE_RELOAD is set to True. Loading data files.')
    else:
        raise RuntimeError(
            'Cache file exists, but FORCE_RELOAD is not set to True. It should '
            'not be possible to reach this point in the code in this case, '
            'please check for errors in the code.'
        )
    data_dict: dict[str, pyam.IamDataFrame | Exception |
                        dict[str, pyam.IamDataFrame | Exception]] = dict()
    _p: Path
    _idf: pyam.IamDataFrame | Exception | dict[str, pyam.IamDataFrame | Exception]
    relpaths: dict[Path, Path] = {
        _p: _p.relative_to(data_root)
        for _p in data_root.glob('**/*.xlsx')
    }
    _relpath: Path
    _abspath: Path
    for _abspath, _relpath in relpaths.items():
        _idf = open_multisheet_iamc(_abspath)
        data_dict[str(_relpath)] = _idf
    if write_cache:
        if cache_file.exists():
            raise FileExistsError(
                f'Cache file {cache_file} already exists. Please delete it '
                'before writing a new cache file, or set a different name/path '
                'in the code.'
            )
        with open(cache_file, 'wb') as f:
            pickle.dump(data_dict, f)
        # Calculate the hash of the pickle file and print it.
        with open(cache_file, 'rb') as f:
            cache_file_content = f.read()
            cache_file_hash = hashlib.md5(cache_file_content).hexdigest()
            print(
                f'Cache file written to {str(cache_file)} with MD5 hash '
                f'{cache_file_hash}'
            )

# %%
# Flatten the dict by inserting an @ sign in front of tab names for files that
# have multiple tabs
flat_data_dict: dict[str, pyam.IamDataFrame | Exception] = dict()
_pathkey: str
for _pathkey, _idf in data_dict.items():
    if isinstance(_idf, dict):
        for _tabname, _idf_tab in _idf.items():
            flat_data_dict[f'{_relpath}@{_tabname}'] = _idf_tab
    else:
        flat_data_dict[_pathkey] = _idf

# %%
# Get only the items that are IamDataFrames
flat_data_dict_iamdfs: dict[str, pyam.IamDataFrame] = {
    _relpath: _idf for _relpath, _idf in flat_data_dict.items()
    if isinstance(_idf, pyam.IamDataFrame)
}

# %%
# Find all the scenarios in the data files
_datafile: str
file_scenarios: dict[str, list[str]] = {
    _datafile: _idf.scenario for _datafile, _idf in flat_data_dict.items()
    if isinstance(_idf, pyam.IamDataFrame)
}

# %%
# Find all the variables in the data files
file_variables: dict[str, list[str]] = {
    _datafile: _idf.variable for _datafile, _idf in flat_data_dict.items()
    if isinstance(_idf, pyam.IamDataFrame)
}

# %%
# Get a dsd and region processor
dsd: nomenclature.DataStructureDefinition = icnom.get_dsd()
regproc: nomenclature.RegionProcessor = icnom.get_region_processor()

# %%
# Get invalid variable names from all the files, using iamcompact_nomenclature.validation
check_dim: str = 'variable'
invalid_names: dict[str, list[str]] = {
    _pathkey: icnom.validation.get_invalid_names(_idf)[check_dim]
    for _pathkey, _idf in flat_data_dict_iamdfs.items()
}

# %%
# Go the other way: For each invalid variable name, find the files that contain
# it. Create a dict with invalid variable names as keys, and a list of file
# IDs as values.
invalid_var_files: dict[str, list[str]] = {
    _varname: [
        _pathkey for _pathkey, _invalids in invalid_names.items()
        if _varname in _invalids
    ]
    for _varname in sorted(list(set().union(*invalid_names.values())))
}

# %%
# Create a dict with file ids as keys and *all* variable names for each file as
# values
file_allvars: dict[str, list[str]] = {
    _pathkey: _idf.variable
    for _pathkey, _idf in flat_data_dict_iamdfs.items()
}

# %%
# Then create the opposite dict: With variable names as keys, and a list of file
# ids as values
# PT = tp.TypeVar('PT')
# def force_type(x, t: tp.Type[PT]) -> PT:
#    return x

var_files = {
    _varname: [
        _pathkey for _pathkey, _allvars in file_allvars.items()
        if _varname in _allvars
    ]
    for _varname in sorted(list(set().union(*file_allvars.values())))
}
