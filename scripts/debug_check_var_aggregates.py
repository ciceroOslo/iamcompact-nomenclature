"""Debugging `aggregation.check_var_aggregates`"""

# %%
# Imports
from pathlib import Path
import typing as tp

import pyam
import pandas as pd
import nomenclature
import openpyxl

import iamcompact_nomenclature as icnom
from iamcompact_nomenclature.aggregation import check_var_aggregates
from iamcompact_nomenclature import var_utils


# %%
# Create a helper function that takes a function as an argument, and returns
# a modified function which will return the return value of the original
# function if it returns, but catch any exceptions and return them. The function
# should be properly type-annotated, with ParamSpecs to ensure that the
# modified function has the same signature and return type as the original
# function, except that it can return an exception as well as the original
# return type.
PS = tp.ParamSpec('PS')  # Parameter specifiction
RT = tp.TypeVar('RT')  # Return type
def return_exceptions(
    func: tp.Callable[PS, RT]
) -> tp.Callable[PS, RT|Exception]:
    def wrapper(*args: PS.args, **kwargs: PS.kwargs) -> RT|Exception:
        try:
            return func(*args, **kwargs)
        except Exception as e:
            return e
    return wrapper
# def return_exceptions(
#     func: tp.Callable[..., tp.Any]
# ) -> tp.Callable[..., tp.Union[tp.Any, Exception]]:
#     def wrapper(*args, **kwargs) -> tp.Union[tp.Any, Exception]:
#         try:
#             return func(*args, **kwargs)
#         except Exception as e:
#             return e
#     return wrapper

# %%
# Define a function to open IAMC excel files with multiple sheets
def open_multisheet_iamc(path: Path|str) -> \
        pyam.IamDataFrame|Exception|dict[str, pyam.IamDataFrame|Exception]:
    sheetnames: list[str] = openpyxl.open(path).sheetnames
    if len(sheetnames) == 1:
        return return_exceptions(pyam.IamDataFrame)(path)
    else:
        return {
            sheetname: return_exceptions(pyam.IamDataFrame) \
                (path, sheet_name=sheetname)
            for sheetname in sheetnames
        }

# %%
# Get data files, and create a nested dict of IamDataFrames
data_root: Path = Path.home() / 'src' / 'repos' / 'pyam_analyses' / 'data' \
    / 'IAM_COMPACT' / 'studies'
data_files_flat: dict[Path, pyam.IamDataFrame|Exception|dict[str, pyam.IamDataFrame|Exception]] = {
    p: open_multisheet_iamc(p)
    for p in data_root.glob('**/*.xlsx')
}
# data_file: Path = Path.home() / 'src' / 'repos' / 'pyam_analyses' / 'data' \
#     / 'IAM_COMPACT' / 'studies' / 'S01_NECPs' \
#         / 'iam_compact_st1__GCAM_iamc_report.xlsx'
#         # / 'Report_IAM_COMPACT_PROMETHEUS_regions_nolinksv2.xlsx'
# idf: pyam.IamDataFrame = pyam.IamDataFrame(data_file)

# %%
# Get a dsd and region processor
dsd: nomenclature.DataStructureDefinition = icnom.get_dsd()
regproc: nomenclature.RegionProcessor = icnom.get_region_processor()

# %%
# Do a test variable aggregation check
var_res: icnom.aggregation.VarAggregationCheckResult = \
    icnom.aggregation.check_var_aggregates(iamdf=idf, dsd=dsd)

# %%
# Do a test region aggregation check
reg_res: icnom.aggregation.RegionAggregationCheckResult = \
    icnom.aggregation.check_region_aggregates(
        iamdf=idf,
        dsd=dsd,
        processor=regproc,
    )

# %%
# Test aggregation of selected variables
failed_checks_partial: pd.DataFrame|None
aggregation_map_partial: dict[str, list[str]]
failed_checks_partial, aggregation_map_partial = \
    icnom.aggregation.check_var_aggregates_manual(
        iamdf=idf,
        # variables=['Capacity|Electricity']
    )

# %%
# Test aggregation of all variables
failed_checks: pd.DataFrame|None
aggregation_map: dict[str, list[str]]
failed_checks, aggregation_map = check_var_aggregates(
    iamdf=idf,
)

# %%
# Now test region mapping
regmap: nomenclature.RegionProcessor = icnom.get_region_processor()
